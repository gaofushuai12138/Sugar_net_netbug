import os
import requests
from openpyxl import load_workbook

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
}
# url = "https://www.duitang.com/napi/blog/list/by_search/?kw=%s&type=feed&start=%d"
url ='https://www.duitang.com/napi/blog/list/by_search/?kw=%s&start=%d'
params = {
    'type':'feed',
    'include_fields':'top_comments,is_root,source_link,item,buyable,root_id,status,like_count,like_id,sender,album,reply_count,favorite_blog_id',
    '_':'1618222003872'
}
print("====================")
print("1.读取excel\n")
print("2.爬取相册\n")
print("0.退出")
print("====================")
num = input("请输入你要选择的功能:\n")


if int(num) == 1:
    filePath = input("请输入excel文件路径:")
    star_num = input("请选取点赞数排名前n的图片:(输入数字)")
    collot_num = input("请选取收藏数排名前n的图片:(输入数字)")
    wb = load_workbook(filename=filePath)

    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    # 第一个表格的名称
    sheet_first = sheets[0]

    # 获取特定的worksheet
    ws = wb[sheet_first]
    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns
    path2 = 'D://tu'
    def get_url(keyword, page):
        new_url = format(url % (keyword, page))
        return new_url
    s = requests.session()
    s.keep_alive = False
    requests.DEFAULT_RETRIES = 5
    img_arr = []
    for row in rows:
        line = [col.value for col in row]

        for item in line:
            print(item)
            initPage = 0
            stritem = item
            response = requests.get(url=get_url(item,initPage), headers=headers,params=params)
            total = response.json()['data']['total']
            next_page = response.json()['data']['next_start']
            print(response.json())
            for imgitem in response.json()['data']['object_list']:
                print(imgitem)
                img_arr.append(imgitem)
            for i in range(next_page, total, 24):
                response = requests.get(url=get_url(stritem, i), headers=headers,params=params, timeout=5000)
                for items in response.json()['data']['object_list']:
                    print(items['photo']['path'] + '\n')
                    img_arr.append(items)
    img_arr.sort(key=lambda x:x['favorite_count'])
    fav = 1
    filePath = "d://"+stritem+"点赞前"+star_num+"收藏前"+collot_num+"名"
    if os.path.exists(filePath) == False:
        os.makedirs(filePath)

    for img_arr_fav_item in img_arr[-int(collot_num):]:
        fav += 1
        image = requests.get(img_arr_fav_item['photo']['path']).content
        with open(filePath+'/fav'+str(fav)+".jpg",'wb') as fp:
            fp.write(image)

    img_arr.sort(key=lambda x:x['like_count'])
    like = 1
    for img_arr_like_item in img_arr[-int(star_num):]:
        like += 1
        image = requests.get(img_arr_like_item['photo']['path']).content

        with open(filePath+'/lik'+str(like)+'.jpg','wb') as fp:
            fp.write(image)
elif int(num) == 2:
     pass
     # albumurl = "https://www.duitang.com/napi/index/hot/?start=%d"
     # albumNum = 1
     # paramsal = {
     #     "include_fileds":'top_comments,is_root,source_link,item,buyable,root_id,status,like_count,sender,album',
     #     "limit":"24",
     #     "_":"1618232462686"
     # }
     # def getAlbumInfo(id):
     #     return format(albumurl%id)
     # response = requests.get(getAlbumInfo(0),headers=headers,params=paramsal)
     # altotal = response.json()['data']['total']
     # for i in range(0,altotal,24):
     #
     #     print(getAlbumInfo(i))
     #     response = requests.get(getAlbumInfo(i),headers = headers,params=paramsal)
     #     for alitem in response.json()['data']['object_list']:
     #         print(alitem['album']['id'])


else:
    print("程序退出!")










        # imgpath = response.json()['data']['object_list'][0]['photo']['path']
        # list_data = response.json()['data']
        # print(list_data)






